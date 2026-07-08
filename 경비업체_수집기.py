# -*- coding: utf-8 -*-
"""
경비업체 목록 수집기 (네이버 검색 API 기반)
============================================
[중요] 경비업 허가는 시·도경찰청(지방경찰청) 소관이라 지방행정 인허가데이터
(localdata.go.kr)에는 포함되지 않고, 공공데이터포털에는 지역별 "집계 통계"만
공개되어 있어 개별 업체명·주소·전화번호가 담긴 공식 공공데이터가 없습니다.

그래서 이 버전은 네이버 검색(지역검색) API로 지역+키워드 조합을 여러 번 검색해
업체 정보를 모으는 방식으로 동작합니다. 공식 등록 전수는 아니지만, 실제 영업 중이고
네이버에 노출된 업체들을 폭넓게 수집합니다.

사전 준비
---------
1. https://developers.naver.com 에서 애플리케이션 등록 → '검색' API 사용 설정
2. 발급받은 Client ID / Client Secret 을 아래 NAVER_CLIENT_ID / NAVER_CLIENT_SECRET 에 입력

실행
----
pip install requests beautifulsoup4 openpyxl
python 경비업체_수집기.py
"""

import re
import time
import datetime
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

# ============================================================
# 설정
# ============================================================

# NAVER_CLIENT_ID / NAVER_CLIENT_SECRET 은 환경변수에서 읽습니다.
# - 로컬 실행: 아래 os.environ.get(...) 의 두 번째 인자(기본값) 자리에 직접 키를 넣어도 됩니다.
#   예) NAVER_CLIENT_ID = os.environ.get("NAVER_CLIENT_ID", "발급받은_Client_ID")
# - GitHub Actions 실행: 저장소 Secrets에 등록된 값이 자동으로 주입됩니다. (코드 수정 불필요)
NAVER_CLIENT_ID = os.environ.get("NAVER_CLIENT_ID", "")
NAVER_CLIENT_SECRET = os.environ.get("NAVER_CLIENT_SECRET", "")

# 수집 대상 지역: {시/도: [구/군, ...]} 형태로 관리
# 전국 17개 시·도, 229개 시·군·구 전체 (일부만 원하면 아래 딕셔너리에서 지워도 됨)
TARGET_REGIONS = {
    "서울": [
        "종로구", "중구", "용산구", "성동구", "광진구", "동대문구", "중랑구",
        "성북구", "강북구", "도봉구", "노원구", "은평구", "서대문구", "마포구",
        "양천구", "강서구", "구로구", "금천구", "영등포구", "동작구", "관악구",
        "서초구", "강남구", "송파구", "강동구",
    ],
    "부산": [
        "강서구", "금정구", "남구", "동구", "동래구", "부산진구", "북구",
        "사상구", "사하구", "서구", "수영구", "연제구", "영도구", "중구",
        "해운대구", "기장군",
    ],
    "대구": ["남구", "달서구", "동구", "북구", "서구", "수성구", "중구", "달성군"],
    "인천": [
        "계양구", "남동구", "동구", "미추홀구", "부평구", "서구", "연수구",
        "중구", "강화군", "옹진군",
    ],
    "광주": ["광산구", "남구", "동구", "북구", "서구"],
    "대전": ["대덕구", "동구", "서구", "유성구", "중구"],
    "울산": ["남구", "동구", "북구", "중구", "울주군"],
    "세종": [""],
    "경기": [
        "수원시", "성남시", "의정부시", "안양시", "부천시", "광명시", "평택시",
        "동두천시", "안산시", "고양시", "과천시", "구리시", "남양주시", "오산시",
        "시흥시", "군포시", "의왕시", "하남시", "용인시", "파주시", "이천시",
        "안성시", "김포시", "화성시", "광주시", "양주시", "포천시", "여주시",
        "연천군", "가평군", "양평군",
    ],
    "강원": [
        "춘천시", "원주시", "강릉시", "동해시", "태백시", "속초시", "삼척시",
        "홍천군", "횡성군", "영월군", "평창군", "정선군", "철원군", "화천군",
        "양구군", "인제군", "고성군", "양양군",
    ],
    "충북": [
        "청주시", "충주시", "제천시", "보은군", "옥천군", "영동군", "증평군",
        "진천군", "괴산군", "음성군", "단양군",
    ],
    "충남": [
        "천안시", "공주시", "보령시", "아산시", "서산시", "논산시", "계룡시",
        "당진시", "금산군", "부여군", "서천군", "청양군", "홍성군", "예산군",
        "태안군",
    ],
    "전북": [
        "전주시", "군산시", "익산시", "정읍시", "남원시", "김제시", "완주군",
        "진안군", "무주군", "장수군", "임실군", "순창군", "고창군", "부안군",
    ],
    "전남": [
        "목포시", "여수시", "순천시", "나주시", "광양시", "담양군", "곡성군",
        "구례군", "고흥군", "보성군", "화순군", "장흥군", "강진군", "해남군",
        "영암군", "무안군", "함평군", "영광군", "장성군", "완도군", "진도군",
        "신안군",
    ],
    "경북": [
        "포항시", "경주시", "김천시", "안동시", "구미시", "영주시", "영천시",
        "상주시", "문경시", "경산시", "의성군", "청송군", "영양군", "영덕군",
        "청도군", "고령군", "성주군", "칠곡군", "예천군", "봉화군", "울진군",
        "울릉군",
    ],
    "경남": [
        "창원시", "진주시", "통영시", "사천시", "김해시", "밀양시", "거제시",
        "양산시", "의령군", "함안군", "창녕군", "고성군", "남해군", "하동군",
        "산청군", "함양군", "거창군", "합천군",
    ],
    "제주": ["제주시", "서귀포시"],
}

# 지역명 뒤에 붙일 검색 키워드 (여러 개를 조합해 커버리지를 넓힘)
SEARCH_KEYWORDS = ["경비업체", "시설경비", "보안업체"]

# 결과 파일 저장 폴더 (GitHub Actions에서 커밋 대상 폴더로도 사용)
OUTPUT_DIR = "data"

FETCH_HOMEPAGE_FOR_EMAIL = True
REQUEST_TIMEOUT = 5
REQUEST_DELAY = 0.4

EMAIL_REGEX = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
}
TAG_STRIP = re.compile(r"<[^>]+>")


# ============================================================
# 1. 네이버 지역검색으로 업체 수집
# ============================================================
def search_naver_local(query, display=5):
    url = "https://openapi.naver.com/v1/search/local.json"
    headers = {
        "X-Naver-Client-Id": NAVER_CLIENT_ID,
        "X-Naver-Client-Secret": NAVER_CLIENT_SECRET,
    }
    try:
        resp = requests.get(
            url, headers=headers,
            params={"query": query, "display": display}, timeout=REQUEST_TIMEOUT
        )
        if resp.status_code == 200:
            return resp.json().get("items", [])
        else:
            print(f"  [경고] 네이버 API 오류({resp.status_code}): {query}")
            return []
    except requests.RequestException as e:
        print(f"  [경고] 요청 실패: {query} ({e})")
        return []


def clean(text):
    return TAG_STRIP.sub("", text or "").strip()


def collect_all():
    if not NAVER_CLIENT_ID or not NAVER_CLIENT_SECRET:
        print("[오류] NAVER_CLIENT_ID / NAVER_CLIENT_SECRET 환경변수가 설정되지 않았습니다.")
        print("로컬 실행 시: set NAVER_CLIENT_ID=발급받은값 (Windows) 후 실행")
        print("GitHub Actions 실행 시: 저장소 Settings > Secrets에 등록되어 있는지 확인하세요.")
        return []

    rows = []
    total_queries = sum(len(v) for v in TARGET_REGIONS.values()) * len(SEARCH_KEYWORDS)
    done = 0
    for city, districts in TARGET_REGIONS.items():
        for district in districts:
            for keyword in SEARCH_KEYWORDS:
                query = " ".join(filter(None, [city, district, keyword]))
                items = search_naver_local(query, display=5)
                done += 1
                print(f"[검색 {done}/{total_queries}] {query} -> {len(items)}건")
                for item in items:
                    name = clean(item.get("title", ""))
                    address = item.get("roadAddress") or item.get("address", "")
                    phone = item.get("telephone", "")
                    link = item.get("link", "")
                    homepage = link if link and "naver.com" not in link else ""
                    if not name:
                        continue
                    rows.append({
                        "업체명": name,
                        "주소": address,
                        "전화번호": phone,
                        "홈페이지": homepage,
                        "이메일": "",
                    })
                time.sleep(REQUEST_DELAY)
    print(f"[정보] 네이버 검색으로 총 {len(rows)}건(중복 포함) 수집")
    return rows


# ============================================================
# 2. 홈페이지에서 공개 이메일 추출 (best-effort)
# ============================================================
def extract_email_from_homepage(url):
    try:
        resp = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
        resp.encoding = resp.apparent_encoding
        soup = BeautifulSoup(resp.text, "html.parser")
        text = soup.get_text(" ")
        emails = EMAIL_REGEX.findall(text)
        emails = [e for e in emails if not e.lower().endswith((".png", ".jpg", ".gif", ".svg"))]
        return emails[0] if emails else ""
    except requests.RequestException:
        return ""


def enrich_with_email(rows):
    if not FETCH_HOMEPAGE_FOR_EMAIL:
        return rows
    targets = [r for r in rows if r["홈페이지"]]
    for i, r in enumerate(targets, start=1):
        r["이메일"] = extract_email_from_homepage(r["홈페이지"])
        print(f"  [이메일 수집 {i}/{len(targets)}] {r['업체명']} -> {'찾음' if r['이메일'] else '없음'}")
        time.sleep(REQUEST_DELAY)
    return rows


# ============================================================
# 3. 중복 제거
# ============================================================
def dedupe(rows):
    seen = {}
    for r in rows:
        key = (r["업체명"].replace(" ", ""), r["전화번호"].replace("-", ""))
        if key not in seen:
            seen[key] = r
        else:
            existing = seen[key]
            for field in ("주소", "전화번호", "홈페이지", "이메일"):
                if not existing.get(field) and r.get(field):
                    existing[field] = r[field]
    result = list(seen.values())
    print(f"[정보] 중복 제거 후 {len(result)}건")
    return result


# ============================================================
# 4. xlsx 저장 (재실행 시 기존 파일과 병합/업데이트)
# ============================================================
COLUMNS = ["업체명", "주소", "전화번호", "이메일", "홈페이지", "최초수집일", "최종수정일"]


def load_existing(path):
    existing = {}
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(headers, row))
            key = (str(d.get("업체명", "")).replace(" ", ""),
                   str(d.get("전화번호", "")).replace("-", ""))
            existing[key] = d
    return existing


def save_xlsx(rows, path):
    today = datetime.date.today().isoformat()
    existing = load_existing(path)

    merged = {k: dict(v) for k, v in existing.items()}

    for r in rows:
        key = (r["업체명"].replace(" ", ""), r["전화번호"].replace("-", ""))
        if key in merged:
            old = merged[key]
            changed = False
            for field in ("주소", "전화번호", "이메일", "홈페이지"):
                new_val = r.get(field, "")
                if new_val and new_val != old.get(field, ""):
                    old[field] = new_val
                    changed = True
            old["최종수정일"] = today if changed else old.get("최종수정일", today)
        else:
            merged[key] = {
                "업체명": r["업체명"],
                "주소": r.get("주소", ""),
                "전화번호": r.get("전화번호", ""),
                "이메일": r.get("이메일", ""),
                "홈페이지": r.get("홈페이지", ""),
                "최초수집일": today,
                "최종수정일": today,
            }

    wb = Workbook()
    ws = wb.active
    ws.title = "경비업체목록"

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, d in enumerate(
        sorted(merged.values(), key=lambda x: x.get("업체명", "")), start=2
    ):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=d.get(col_name, ""))

    widths = [24, 40, 16, 26, 30, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    wb.save(path)
    print(f"[완료] {path} 저장 (총 {len(merged)}건)")


# ============================================================
# 실행
# ============================================================
def build_output_filename():
    if len(TARGET_REGIONS) == 1:
        city = list(TARGET_REGIONS.keys())[0]
        name = f"{city}_경비업체_목록.xlsx"
    else:
        name = "전국_경비업체_목록.xlsx"
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    return os.path.join(OUTPUT_DIR, name)


def main():
    rows = collect_all()
    if not rows:
        return
    rows = enrich_with_email(rows)
    rows = dedupe(rows)
    save_xlsx(rows, build_output_filename())


if __name__ == "__main__":
    main()
